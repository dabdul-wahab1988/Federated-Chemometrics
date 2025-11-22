from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Dict, List, Tuple, Optional, Any

import sys
import pandas as pd
import numpy as np

# Ensure the project src/ tree is importable when running from a checkout
try:
    import fedchem  # type: ignore  # noqa: F401
except Exception:
    _proj_root = Path(__file__).resolve().parents[1]
    _src_dir = _proj_root / "src"
    if _src_dir.exists():
        sys.path.insert(0, str(_src_dir))

from fedchem.results import ensure_results_tree
from fedchem.utils.config import (
    load_config,
    get_experimental_sites,
    get_data_config,
    get_instrument_to_site_map,
)


@dataclass
class SiteMeta:
    instrument_id: str
    manufacturer: str
    role: Optional[str]
    enabled: bool
    cal_rows: int
    val_rows: int
    test_rows: int


def load_results_tree(root: Path):
    """
    Wrapper around ensure_results_tree for easy reuse.
    """
    return ensure_results_tree(root)


def load_config_dict(path: Path | str = "config.yaml") -> dict:
    return load_config(Path(path))


def _build_loader_instrument_map(cfg: dict) -> Dict[str, dict]:
    """
    Build a loader-friendly instrument map keyed by multiple aliases.

    This mirrors the logic in scripts/run_real_site_experiment.py so that
    sheet names such as 'CalSetA1' resolve correctly to config entries
    like 'MA_A1'.
    """
    instrument_map = get_instrument_to_site_map(cfg)
    loader_map: Dict[str, dict] = {}
    if not instrument_map:
        return loader_map
    for iid, meta in instrument_map.items():
        loader_map[iid] = meta
        suffix = str(iid).split("_")[-1]
        loader_map.setdefault(suffix, meta)
        calset_key = f"CalSet{suffix}"
        loader_map.setdefault(calset_key, meta)
    return loader_map


def load_site_metadata(
    cfg: dict,
    data_dir: Path,
    *,
    n_wavelengths: Optional[int] = None,
) -> Tuple[Dict[str, SiteMeta], dict]:
    """
    Load IDRC per-instrument metadata (sample counts, mapping to manufacturers).

    Returns a mapping instrument_id -> SiteMeta and a lightweight
    meta dict containing per-instrument row counts.
    """
    from openpyxl import load_workbook  # type: ignore

    instrument_map_full = get_instrument_to_site_map(cfg) or {}
    instrument_meta: Dict[str, SiteMeta] = {}

    # Pre-compute val/test row counts per manufacturer
    manufacturers = {
        str(meta.get("site"))
        for meta in instrument_map_full.values()
        if meta.get("site") is not None
    }
    val_counts: Dict[str, int] = {}
    test_counts: Dict[str, int] = {}
    for mfr in manufacturers:
        val_rows = 0
        test_rows = 0
        val_path = data_dir / mfr / f"Val_{mfr}.xlsx"
        test_path = data_dir / mfr / f"Test_{mfr}.xlsx"
        if val_path.exists():
            wb = load_workbook(val_path, read_only=True)
            ws = wb[wb.sheetnames[0]]
            # subtract header row if present
            val_rows = max(0, ws.max_row - 1)
        if test_path.exists():
            wb = load_workbook(test_path, read_only=True)
            ws = wb[wb.sheetnames[0]]
            test_rows = max(0, ws.max_row - 1)
        val_counts[mfr] = val_rows
        test_counts[mfr] = test_rows

    # Calibration rows per instrument from Cal_<Manufacturer>.xlsx sheets
    cal_workbooks: Dict[str, Any] = {}
    instrument_meta: Dict[str, SiteMeta] = {}
    for instr_id, meta in instrument_map_full.items():
        manufacturer = str(meta.get("site") or "")
        if not manufacturer:
            continue
        # Sheet names follow CalSetA1 / CalSetB2, etc.
        suffix = str(instr_id).split("_")[-1]
        sheet_name = f"CalSet{suffix}"
        wb = cal_workbooks.get(manufacturer)
        if wb is None:
            cal_path = data_dir / manufacturer / f"Cal_{manufacturer}.xlsx"
            if not cal_path.exists():
                continue
            wb = load_workbook(cal_path, read_only=True)
            cal_workbooks[manufacturer] = wb
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        cal_rows = max(0, ws.max_row - 1)
        val_rows = val_counts.get(manufacturer, 0)
        test_rows = test_counts.get(manufacturer, 0)
        instrument_meta[instr_id] = SiteMeta(
            instrument_id=str(instr_id),
            manufacturer=manufacturer,
            role=meta.get("role"),
            enabled=bool(meta.get("enabled", True)),
            cal_rows=cal_rows,
            val_rows=val_rows,
            test_rows=test_rows,
        )

    # Build a minimal meta dict mirroring the structure used elsewhere
    site_summaries = {
        instr_id: {
            "instrument_id": instr_id,
            "backing_site": sm.manufacturer,
            "cal_rows": sm.cal_rows,
            "val_rows": sm.val_rows,
            "test_rows": sm.test_rows,
        }
        for instr_id, sm in instrument_meta.items()
    }
    meta_out: dict = {
        "dataset": "idrc_wheat_shootout",
        "site_summaries": site_summaries,
    }
    return instrument_meta, meta_out


def load_publication_database(pub_db_root: Path) -> Tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    """
    Load publication database CSVs.

    Returns:
      - master_database (complete integrated database)
      - performance_metrics (performance data for all methods)
      - conformal_metrics (conformal prediction coverage)
      - training_dynamics (round-by-round training data)
    """
    csv_dir = pub_db_root / "csv"
    if not csv_dir.exists():
        return pd.DataFrame(), pd.DataFrame(), pd.DataFrame(), pd.DataFrame()
    
    # Find the most recent master database file
    master_files = sorted(csv_dir.glob("master_database_*.csv"))
    perf_files = sorted(csv_dir.glob("performance_metrics_*.csv"))
    conf_files = sorted(csv_dir.glob("conformal_metrics_*.csv"))
    train_files = sorted(csv_dir.glob("training_dynamics_*.csv"))
    
    master_df = pd.read_csv(master_files[-1]) if master_files else pd.DataFrame()
    perf_df = pd.read_csv(perf_files[-1]) if perf_files else pd.DataFrame()
    conf_df = pd.read_csv(conf_files[-1]) if conf_files else pd.DataFrame()
    train_df = pd.read_csv(train_files[-1]) if train_files else pd.DataFrame()
    
    # Ensure compatibility: many upstream scripts expect a 'site_id' column in methods/perf df
    if perf_df is not None and 'site_id' not in perf_df.columns and 'instrument_code' in perf_df.columns:
        perf_df['site_id'] = perf_df['instrument_code']

    return master_df, perf_df, conf_df, train_df


def load_aggregated_frames(results_root: Path, use_publication_db: bool = True) -> Tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    """
    Load aggregated CSVs from publication database (default) or legacy aggregated files.

    Returns:
      - methods_summary_all (or performance_metrics from pub DB)
      - metrics_summary (empty for pub DB compatibility)
      - raw_records (or master_database from pub DB)
    """
    # Use publication database by default
    if use_publication_db:
        pub_db_path = results_root / "publication_database"
        if pub_db_path.exists():
            master_df, perf_df, _, _ = load_publication_database(pub_db_path)
            if not master_df.empty:
                # Map publication DB column names to legacy format for compatibility
                master_mapped = master_df.copy()

                # Drop aggregate-only rows that should not appear in plots/tables
                master_mapped = master_mapped[master_mapped["method"] != "mean_rmse"].copy()
                
                # Map column names from publication DB to expected legacy format
                column_mapping = {
                    'transfer_k': 'design_factors.Transfer_Samples',
                    'dp_epsilon': 'design_factors.DP_Target_Eps',
                    # Keep instrument_code as-is (used by both formats)
                    'rmsep': 'metrics.RMSEP',
                    'r2': 'metrics.R2',
                    'bytes_mb': 'runtime.total_bytes_mb',
                    'round_time_sec': 'metrics.Round_Time',
                    'bytes_sent': 'metrics.Bytes_Sent',
                    'bytes_received': 'metrics.Bytes_Received',
                    'total_bytes': 'metrics.Total_Bytes',
                    'wall_time_sec': 'runtime.wall_time_total_sec',
                    'num_rounds': 'metrics.Rounds'
                }
                
                # Only rename columns that exist
                rename_dict = {k: v for k, v in column_mapping.items() if k in master_mapped.columns}
                master_mapped = master_mapped.rename(columns=rename_dict)
                
                # Add run_label for compatibility (objective_1 is the main experiment)
                if 'run_label' not in master_mapped.columns:
                    master_mapped['run_label'] = 'objective_1'
                else:
                    master_mapped['run_label'] = master_mapped['run_label'].fillna('objective_1')
                
                # Convert dp_epsilon to consistent format (all strings for sorting compatibility)
                if 'design_factors.DP_Target_Eps' in master_mapped.columns:
                    def convert_epsilon(x):
                        if isinstance(x, float):
                            if np.isinf(x):
                                return 'inf'
                            else:
                                return str(float(x))  # Convert float to string to avoid mixed-type sorting issues
                        elif isinstance(x, str):
                            return x
                        else:
                            return str(x)
                    master_mapped['design_factors.DP_Target_Eps'] = master_mapped['design_factors.DP_Target_Eps'].apply(convert_epsilon)

                # Ensure transfer_k is numeric (int) where possible
                if 'design_factors.Transfer_Samples' in master_mapped.columns:
                    master_mapped['design_factors.Transfer_Samples'] = master_mapped['design_factors.Transfer_Samples'].apply(
                        lambda v: int(v) if pd.notna(v) else v
                    )

                # Normalize baseline methods to be non-DP and tagged as "objective_1"
                baseline_methods = {"Centralized_PLS", "Site_Specific", "PDS", "SBC"}
                if 'method' in master_mapped.columns:
                    # Force pooled baseline eps=inf to avoid misleading DP sweeps
                    if 'design_factors.DP_Target_Eps' in master_mapped.columns:
                        mask_base = master_mapped['method'] == 'Centralized_PLS'
                        master_mapped.loc[mask_base, 'design_factors.DP_Target_Eps'] = 'inf'

                    # Re-classify categories for baselines
                    if 'category' in master_mapped.columns:
                        master_mapped.loc[master_mapped['method'].isin(baseline_methods), 'category'] = 'Classical'

                    # Mark baselines as not using DP when column exists
                    if 'uses_dp' in master_mapped.columns:
                        master_mapped.loc[master_mapped['method'].isin(baseline_methods), 'uses_dp'] = False

                # Remove duplicate rows introduced by epsilon normalization on pooled baseline
                dedup_cols = [c for c in ["method", "instrument_code", "design_factors.Transfer_Samples", "design_factors.DP_Target_Eps"] if c in master_mapped.columns]
                if dedup_cols:
                    master_mapped = master_mapped.drop_duplicates(subset=dedup_cols, keep="first")
                
                # Return in format: (methods_summary, metrics_summary, raw_records)
                return perf_df, pd.DataFrame(), master_mapped
    
    # Fallback to legacy aggregated files
    tree = ensure_results_tree(results_root)
    agg_dir = tree.aggregated
    methods_path = agg_dir / "methods_summary_all.csv"
    metrics_path = agg_dir / "metrics_summary.csv"
    raw_path = agg_dir / "raw_records.csv"

    methods_df = pd.read_csv(methods_path) if methods_path.exists() else pd.DataFrame()
    metrics_df = pd.read_csv(metrics_path) if metrics_path.exists() else pd.DataFrame()
    raw_df = pd.read_csv(raw_path) if raw_path.exists() else pd.DataFrame()
    return methods_df, metrics_df, raw_df


def find_transfer_k_dirs(archive_root: Path) -> List[int]:
    """
    Discover available transfer-sample budgets from the archive tree.

    Looks for subdirectories named `transfer_k_<k>` and returns the
    available k values as integers.
    """
    ks: List[int] = []
    if not archive_root.exists():
        return ks
    for child in archive_root.iterdir():
        if not child.is_dir():
            continue
        name = child.name
        if name.startswith("transfer_k_"):
            token = name.split("_")[-1]
            try:
                ks.append(int(token))
            except Exception:
                continue
    return sorted(sorted(set(ks)))
